from azure.kusto.data.request import KustoClient, KustoConnectionStringBuilder
from azure.kusto.data.helpers import dataframe_from_result_table
import openpyxl
import path_mbr

def authenticate_kusto(kusto_cluster):
    tenant_id = '72f988bf-86f1-41af-91ab-2d7cd011db47'
    KCSB = KustoConnectionStringBuilder.with_aad_device_authentication(kusto_cluster)
    KCSB.authority_id = tenant_id
    return KustoClient(KCSB)

def query_kusto(client, database, query):
    return client.execute(database, query)

cga_cluster = 'https://cgadataout.kusto.windows.net'
cga_client = authenticate_kusto(cga_cluster)
ls=[cga_client, "Docs All", "Learn"]
kusto_database = 'DevRelWorkArea'

def get_query_data_product_docs(client, area, database):
    kusto_query_1 = open(path_mbr.query_path_product_1,'r').read().format(area)
    result_1 = query_kusto(client, database, kusto_query_1)
    df_1 = dataframe_from_result_table(result_1.primary_results[0])
    kusto_query_2 = open(path_mbr.query_path_product_2, 'r').read().format(area)
    result_2 = query_kusto(client, database, kusto_query_2)
    df_2 = dataframe_from_result_table(result_2.primary_results[0])
    kusto_query_3 = open(path_mbr.query_path_product_3, 'r').read().format(area)
    result_3 = query_kusto(client, database, kusto_query_3)
    df_3 = dataframe_from_result_table(result_3.primary_results[0])
    kusto_query_4 = open(path_mbr.query_path_product_4, 'r').read().format(area)
    result_4 = query_kusto(client, database, kusto_query_4)
    df_4 = dataframe_from_result_table(result_4.primary_results[0])
    kusto_query_5 = open(path_mbr.query_path_product_5, 'r').read().format(area)
    result_5 = query_kusto(client, database, kusto_query_5)
    df_5 = dataframe_from_result_table(result_5.primary_results[0])
    kusto_query_6 = open(path_mbr.query_path_product_6, 'r').read().format(area)
    result_6 = query_kusto(client, database, kusto_query_6)
    df_6 = dataframe_from_result_table(result_6.primary_results[0])
    kusto_query_7 = open(path_mbr.query_path_product_7, 'r').read()
    result_7 = query_kusto(client, database, kusto_query_7)
    df_7 = dataframe_from_result_table(result_7.primary_results[0])
    kusto_query_8 = open(path_mbr.query_path_product_8, 'r').read().format(area)
    result_8 = query_kusto(client, database, kusto_query_8)
    df_8 = dataframe_from_result_table(result_8.primary_results[0])
    kusto_query_9 = open(path_mbr.query_path_product_9, 'r').read().format(area, area)
    result_9 = query_kusto(client, database, kusto_query_9)
    df_9 = dataframe_from_result_table(result_9.primary_results[0])
    kusto_query_10 = open(path_mbr.query_path_product_10, 'r').read().format(area)
    result_10 = query_kusto(client, database, kusto_query_10)
    df_10 = dataframe_from_result_table(result_10.primary_results[0])
    return df_1,df_2,df_3,df_4,df_5,df_6,df_7,df_8,df_9,df_10

def get_query_data_product_learn(client, area, database):
    kusto_query_1 = open(path_mbr.query_path_product_1,'r').read().format(area)
    result_1 = query_kusto(client, database, kusto_query_1)
    df_1 = dataframe_from_result_table(result_1.primary_results[0])
    kusto_query_2 = open(path_mbr.query_path_product_2_learn, 'r').read().format(area)
    result_2 = query_kusto(client, database, kusto_query_2)
    df_2 = dataframe_from_result_table(result_2.primary_results[0])
    kusto_query_3 = open(path_mbr.query_path_product_3_learn, 'r').read().format(area)
    result_3 = query_kusto(client, database, kusto_query_3)
    df_3 = dataframe_from_result_table(result_3.primary_results[0])
    kusto_query_4 = open(path_mbr.query_path_product_4_learn, 'r').read().format(area)
    result_4 = query_kusto(client, database, kusto_query_4)
    df_4 = dataframe_from_result_table(result_4.primary_results[0])
    kusto_query_5 = open(path_mbr.query_path_product_5_learn, 'r').read().format(area)
    result_5 = query_kusto(client, database, kusto_query_5)
    df_5 = dataframe_from_result_table(result_5.primary_results[0])
    return df_1,df_2,df_3,df_4,df_5

content_4_1,content_4_2,content_4_3,content_4_4,content_4_5,\
content_4_6,content_4_7,content_4_8,content_4_9,content_4_10 = get_query_data_product_docs(ls[0], ls[1],kusto_database)

content_5_1,content_5_2,content_5_3,content_5_4,content_5_5 = get_query_data_product_learn(ls[0], ls[2],kusto_database)

print(content_4_1,content_4_2,content_4_3,content_4_4,content_4_5,
content_4_6,content_4_7,content_4_8,content_4_9,content_4_10)

print(content_5_1,content_5_2,content_5_3,content_5_4,content_5_5)

wb_obj = openpyxl.load_workbook(path_mbr.path_mbr)
sheet_x_2=wb_obj['PRODUCT']

i = 0
j = 0
x = 9
y = 23
for i in range(14):
    sheet_x_2.cell(x, y, content_4_1.iloc[i][j]) #9行表头
    sheet_x_2.cell(x + 2, y, content_4_1.iloc[i][j + 1]) #11行
    sheet_x_2.cell(x + 3, y, content_4_2.iloc[i][j + 1]) #12行
    sheet_x_2.cell(x + 4, y, content_4_3.iloc[i][j + 1]) #13行
    sheet_x_2.cell(x + 5, y, content_4_4.iloc[i][j + 1]) #14行
    sheet_x_2.cell(x + 6, y, content_4_5.iloc[i][j + 1]) #15行
    sheet_x_2.cell(x + 7, y, content_4_6.iloc[i][j + 1]) #16
    sheet_x_2.cell(x + 9, y, content_4_7.iloc[i][j + 1]) #18
    sheet_x_2.cell(x + 10, y, content_4_8.iloc[i][j + 1]) #19
    sheet_x_2.cell(x + 11, y, content_4_9.iloc[i][j + 1]) #20
    sheet_x_2.cell(x + 13, y, content_4_10.iloc[i][j + 1]) #20
    sheet_x_2.cell(x + 19, y, content_5_1.iloc[i][j + 1]) #28行
    sheet_x_2.cell(x + 20, y, content_5_2.iloc[i][j + 1]) #29行
    sheet_x_2.cell(x + 21, y, content_5_3.iloc[i][j + 1]) #30行
    sheet_x_2.cell(x + 22, y, content_5_4.iloc[i][j + 1]) #31行
    sheet_x_2.cell(x + 24, y, content_5_5.iloc[i][j + 1]) #33行
    y += -1
    i += 1

wb_obj.save(path_mbr.path_mbr)

wb_obj.close()