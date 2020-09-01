from azure.kusto.data.request import KustoClient, KustoConnectionStringBuilder
from azure.kusto.data.helpers import dataframe_from_result_table
import openpyxl
import path_mbr

def authenticate_kusto(kusto_cluster):
    tenant_id = '72f988bf-86f1-41af-91ab-2d7cd011db47'
    KCSB = KustoConnectionStringBuilder.with_aad_device_authentication(kusto_cluster)
    KCSB.authority_id = tenant_id
    return KustoClient(KCSB)

def query_kusto(client, database, query):
    return client.execute(database, query)

cga_cluster = 'https://cgadataout.kusto.windows.net'
cga_client = authenticate_kusto(cga_cluster)
ls=[cga_client]
kusto_database = 'DevRelWorkArea'

def get_query_data_content_operations(client,database):
    kusto_query_1 = open(path_mbr.query_path_content_operations_1, 'r').read()
    result_1 = query_kusto(client, database, kusto_query_1)
    df_1 = dataframe_from_result_table(result_1.primary_results[0])
    kusto_query_2 = open(path_mbr.query_path_content_operations_2, 'r').read()
    result_2 = query_kusto(client, database, kusto_query_2)
    df_2 = dataframe_from_result_table(result_2.primary_results[0])
    kusto_query_3 = open(path_mbr.query_path_content_operations_3, 'r').read()
    result_3 = query_kusto(client, database, kusto_query_3)
    df_3 = dataframe_from_result_table(result_3.primary_results[0])
    kusto_query_4 = open(path_mbr.query_path_content_operations_4, 'r').read()
    result_4 = query_kusto(client, database, kusto_query_4)
    df_4 = dataframe_from_result_table(result_4.primary_results[0])
    kusto_query_5 = open(path_mbr.query_path_content_operations_5, 'r').read()
    result_5 = query_kusto(client, database, kusto_query_5)
    df_5 = dataframe_from_result_table(result_5.primary_results[0])
    kusto_query_6 = open(path_mbr.query_path_content_operations_6, 'r').read()
    result_6 = query_kusto(client, database, kusto_query_6)
    df_6 = dataframe_from_result_table(result_6.primary_results[0])
    kusto_query_8 = open(path_mbr.query_path_content_operations_8, 'r').read()
    result_8 = query_kusto(client, database, kusto_query_8)
    df_8 = dataframe_from_result_table(result_8.primary_results[0])
    return df_1, df_2, df_3, df_4, df_5, df_6, df_8

content_6_1,content_6_2,content_6_3,content_6_4,content_6_5,\
content_6_6,content_6_8 = get_query_data_content_operations(ls[0], kusto_database)

# Change values in Excel
wb_obj = openpyxl.load_workbook(path_mbr.path_mbr)
sheet_x_3=wb_obj['CONTENT_OPERATIONS']

i = 0
j = 0
x = 9
y = 23
for i in range(14):
    sheet_x_3.cell(x + 1, y, content_6_1.iloc[i][j + 1]) #10行
    sheet_x_3.cell(x + 2, y, content_6_2.iloc[i][j + 1]) #11行
    sheet_x_3.cell(x + 7, y, content_6_3.iloc[i][j + 1]) #16行
    sheet_x_3.cell(x + 10, y, content_6_4.iloc[i][j + 1]) #19行
    sheet_x_3.cell(x + 11, y, content_6_5.iloc[i][j + 1]) #20行
    sheet_x_3.cell(x + 13, y, content_6_6.iloc[i][j + 1]) #22
    sheet_x_3.cell(x + 15, y, content_6_8.iloc[i][j + 1]) #24
    y += -1
    i += 1

wb_obj.save(path_mbr.path_mbr)