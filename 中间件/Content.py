from azure.kusto.data.request import KustoClient, KustoConnectionStringBuilder
from azure.kusto.data.helpers import dataframe_from_result_table
import openpyxl
import path_mbr

def authenticate_kusto(kusto_cluster):
    tenant_id = '72f988bf-86f1-41af-91ab-2d7cd011db47'
    KCSB = KustoConnectionStringBuilder.with_aad_device_authentication(kusto_cluster)
    KCSB.authority_id = tenant_id
    return KustoClient(KCSB)

def query_kusto(client, database, query):
    return client.execute(database, query)

cga_cluster = 'https://cgadataout.kusto.windows.net'
cga_client = authenticate_kusto(cga_cluster)
ls=[cga_client, "Documentation & Reference", "Azure Architecture Center", "Learn"]
kusto_database = 'DevRelWorkArea'

def get_page_views(client, area, database):
    kusto_query_1 = open(path_mbr.query_path_1,'r').read().format(area)
    result_1 = query_kusto(client, database, kusto_query_1)
    df_1 = dataframe_from_result_table(result_1.primary_results[0])
    kusto_query_2 = open(path_mbr.query_path_2,'r').read().format(area)
    result_2 = query_kusto(client, database, kusto_query_2)
    df_2 = dataframe_from_result_table(result_2.primary_results[0])
    kusto_query_3 = open(path_mbr.query_path_3,'r').read().format(area)
    result_3 = query_kusto(client, database, kusto_query_3)
    df_3 = dataframe_from_result_table(result_3.primary_results[0])
    kusto_query_4 = open(path_mbr.query_path_4,'r').read().format(area)
    result_4 = query_kusto(client, database, kusto_query_4)
    df_4 = dataframe_from_result_table(result_4.primary_results[0])
    kusto_query_5 = open(path_mbr.query_path_5,'r').read().format(area)
    result_5 = query_kusto(client, database, kusto_query_5)
    df_5 = dataframe_from_result_table(result_5.primary_results[0])
    kusto_query_6 = open(path_mbr.query_path_6,'r').read().format(area)
    result_6 = query_kusto(client, database, kusto_query_6)
    df_6 = dataframe_from_result_table(result_6.primary_results[0])
    kusto_query_7 = open(path_mbr.query_path_7,'r').read().format(area)
    result_7 = query_kusto(client, database, kusto_query_7)
    df_7 = dataframe_from_result_table(result_7.primary_results[0])
    kusto_query_8 = open(path_mbr.query_path_8,'r').read().format(area)
    result_8 = query_kusto(client, database, kusto_query_8)
    df_8 = dataframe_from_result_table(result_8.primary_results[0])
    kusto_query_9 = open(path_mbr.query_path_9,'r').read().format(area)
    result_9 = query_kusto(client, database, kusto_query_9)
    df_9 = dataframe_from_result_table(result_9.primary_results[0])
    kusto_query_10 = open(path_mbr.query_path_10,'r').read().format(area)
    result_10 = query_kusto(client, database, kusto_query_10)
    df_10 = dataframe_from_result_table(result_10.primary_results[0])
    kusto_query_11 = open(path_mbr.query_path_11,'r').read().format(area)
    result_11 = query_kusto(client, database, kusto_query_11)
    df_11 = dataframe_from_result_table(result_11.primary_results[0])
    kusto_query_12 = open(path_mbr.query_path_12,'r').read().format(area)
    result_12 = query_kusto(client, database, kusto_query_12)
    df_12 = dataframe_from_result_table(result_12.primary_results[0])
    kusto_query_13 = open(path_mbr.query_path_13,'r').read().format(area)
    result_13 = query_kusto(client, database, kusto_query_13)
    df_13 = dataframe_from_result_table(result_13.primary_results[0])
    kusto_query_14 = open(path_mbr.query_path_14,'r').read().format(area)
    result_14 = query_kusto(client, database, kusto_query_14)
    df_14 = dataframe_from_result_table(result_14.primary_results[0])
    return df_1,df_2,df_3,df_4,df_5,df_6,df_7,df_8,df_9,df_10,df_11,df_12,df_13,df_14

def get_kusto_query_data_learn(client, area, database):
    kusto_query_1 = open(path_mbr.query_path_1,'r').read().format(area)
    result_1 = query_kusto(client, database, kusto_query_1)
    df_1 = dataframe_from_result_table(result_1.primary_results[0])
    kusto_query_2 = open(path_mbr.query_path_2,'r').read().format(area)
    result_2 = query_kusto(client, database, kusto_query_2)
    df_2 = dataframe_from_result_table(result_2.primary_results[0])
    kusto_query_3 = open(path_mbr.query_path_3,'r').read().format(area)
    result_3 = query_kusto(client, database, kusto_query_3)
    df_3 = dataframe_from_result_table(result_3.primary_results[0])
    kusto_query_4 = open(path_mbr.query_path_4,'r').read().format(area)
    result_4 = query_kusto(client, database, kusto_query_4)
    df_4 = dataframe_from_result_table(result_4.primary_results[0])
    kusto_query_5_learn = open(path_mbr.query_path_5_learn, 'r').read().format(area)
    result_5_learn = query_kusto(client, database, kusto_query_5_learn)
    df_5_learn = dataframe_from_result_table(result_5_learn.primary_results[0])
    kusto_query_6 = open(path_mbr.query_path_6,'r').read().format(area)
    result_6 = query_kusto(client, database, kusto_query_6)
    df_6 = dataframe_from_result_table(result_6.primary_results[0])
    kusto_query_7 = open(path_mbr.query_path_7,'r').read().format(area)
    result_7 = query_kusto(client, database, kusto_query_7)
    df_7 = dataframe_from_result_table(result_7.primary_results[0])
    kusto_query_8 = open(path_mbr.query_path_8,'r').read().format(area)
    result_8 = query_kusto(client, database, kusto_query_8)
    df_8 = dataframe_from_result_table(result_8.primary_results[0])
    kusto_query_9 = open(path_mbr.query_path_12,'r').read().format(area)
    result_9 = query_kusto(client, database, kusto_query_9)
    df_9 = dataframe_from_result_table(result_9.primary_results[0])
    kusto_query_10 = open(path_mbr.query_path_5,'r').read().format(area)
    result_10 = query_kusto(client, database, kusto_query_10)
    df_10 = dataframe_from_result_table(result_10.primary_results[0])
    kusto_query_11 = open(path_mbr.query_path_11_learn,'r').read().format(area)
    result_11 = query_kusto(client, database, kusto_query_11)
    df_11 = dataframe_from_result_table(result_11.primary_results[0])
    kusto_query_12 = open(path_mbr.query_path_12_learn,'r').read().format(area)
    result_12 = query_kusto(client, database, kusto_query_12)
    df_12 = dataframe_from_result_table(result_12.primary_results[0])
    kusto_query_13 = open(path_mbr.query_path_13_learn,'r').read().format(area)
    result_13 = query_kusto(client, database, kusto_query_13)
    df_13 = dataframe_from_result_table(result_13.primary_results[0])
    kusto_query_14 = open(path_mbr.query_path_14_learn,'r').read().format(area)
    result_14 = query_kusto(client, database, kusto_query_14)
    df_14 = dataframe_from_result_table(result_14.primary_results[0])
    return df_1,df_2,df_3,df_4,df_5_learn,df_6,df_7,df_8,df_9,df_10,df_11,df_12,df_13,df_14

content_1_1,content_1_2,content_1_3,content_1_4,content_1_5,\
content_1_6,content_1_7,content_1_8,content_1_9,content_1_10,\
content_1_11,content_1_12,content_1_13,content_1_14 = get_page_views(ls[0], ls[1],kusto_database)

content_2_1,content_2_2,content_2_3,content_2_4,content_2_5,\
content_2_6,content_2_7,content_2_8,content_2_9,content_2_10,\
content_2_11,content_2_12,content_2_13,content_2_14 = get_page_views(ls[0], ls[2],kusto_database)

content_3_1,content_3_2,content_3_3,content_3_4,content_3_5,\
content_3_6,content_3_7,content_3_8,content_3_9,content_3_10,\
content_3_11,content_3_12,content_3_13,content_3_14 = get_kusto_query_data_learn(ls[0], ls[3],kusto_database)

wb_obj = openpyxl.load_workbook(path_mbr.path_mbr)
sheet_x = wb_obj['CONTENT']

i = 0
j = 0
x = 9
y = 23
for i in range(14):
    sheet_x.cell(x, y, content_1_1.iloc[i][j]) #9行表头
    sheet_x.cell(x + 1, y, content_1_1.iloc[i][j + 1])
    sheet_x.cell(x + 2, y, content_1_2.iloc[i][j + 1])
    sheet_x.cell(x + 3, y, content_1_3.iloc[i][j + 1])
    sheet_x.cell(x + 4, y, content_1_4.iloc[i][j + 1])
    sheet_x.cell(x + 5, y, content_1_5.iloc[i][j + 1])
    sheet_x.cell(x + 7, y, content_1_6.iloc[i][j + 1])
    sheet_x.cell(x + 8, y, content_1_7.iloc[i][j + 1])
    sheet_x.cell(x + 9, y, content_1_8.iloc[i][j + 1])
    sheet_x.cell(x + 10, y, content_1_9.iloc[i][j + 1])
    sheet_x.cell(x + 11, y, content_1_10.iloc[i][j + 1])
    sheet_x.cell(x + 13, y, content_1_11.iloc[i][j + 1])
    sheet_x.cell(x + 14, y, content_1_12.iloc[i][j + 1])
    sheet_x.cell(x + 15, y, content_1_13.iloc[i][j + 1])
    sheet_x.cell(x + 16, y, content_1_14.iloc[i][j + 1])
    sheet_x.cell(x + 24, y, content_2_1.iloc[i][j + 1])
    sheet_x.cell(x + 25, y, content_2_2.iloc[i][j + 1])
    sheet_x.cell(x + 26, y, content_2_3.iloc[i][j + 1])
    sheet_x.cell(x + 27, y, content_2_4.iloc[i][j + 1])
    sheet_x.cell(x + 28, y, content_2_5.iloc[i][j + 1])
    sheet_x.cell(x + 30, y, content_2_6.iloc[i][j + 1])
    sheet_x.cell(x + 31, y, content_2_7.iloc[i][j + 1])
    sheet_x.cell(x + 32, y, content_2_8.iloc[i][j + 1])
    sheet_x.cell(x + 33, y, content_2_9.iloc[i][j + 1])
    sheet_x.cell(x + 34, y, content_2_10.iloc[i][j + 1])
    sheet_x.cell(x + 36, y, content_2_11.iloc[i][j + 1])
    sheet_x.cell(x + 37, y, content_2_12.iloc[i][j + 1])
    sheet_x.cell(x + 38, y, content_2_13.iloc[i][j + 1])
    sheet_x.cell(x + 39, y, content_2_14.iloc[i][j + 1])
    sheet_x.cell(x + 47, y, content_3_1.iloc[i][j + 1])
    sheet_x.cell(x + 48, y, content_3_2.iloc[i][j + 1])
    sheet_x.cell(x + 49, y, content_3_3.iloc[i][j + 1])
    sheet_x.cell(x + 50, y, content_3_4.iloc[i][j + 1])
    sheet_x.cell(x + 52, y, content_3_5.iloc[i][j + 1])
    sheet_x.cell(x + 53, y, content_3_6.iloc[i][j + 1])
    sheet_x.cell(x + 54, y, content_3_7.iloc[i][j + 1])
    sheet_x.cell(x + 55, y, content_3_8.iloc[i][j + 1])
    sheet_x.cell(x + 57, y, content_3_9.iloc[i][j + 1])
    sheet_x.cell(x + 58, y, content_3_10.iloc[i][j + 1])
    sheet_x.cell(x + 59, y, content_3_11.iloc[i][j + 1])
    sheet_x.cell(x + 61, y, content_3_12.iloc[i][j + 1])
    sheet_x.cell(x + 62, y, content_3_13.iloc[i][j + 1])
    sheet_x.cell(x + 63, y, content_3_14.iloc[i][j + 1])
    y += -1
    i += 1

wb_obj.save(path_mbr.path_mbr)