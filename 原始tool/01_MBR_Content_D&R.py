from azure.kusto.data.request import KustoClient, KustoConnectionStringBuilder
from azure.kusto.data.helpers import dataframe_from_result_table
import openpyxl
import path_mbr

def authenticate_kusto(kusto_cluster):
    tenant_id = '72f988bf-86f1-41af-91ab-2d7cd011db47'
    KCSB = KustoConnectionStringBuilder.with_aad_device_authentication(kusto_cluster)
    KCSB.authority_id = tenant_id
    return KustoClient(KCSB)

def query_kusto(client, database, query):
    return client.execute(database, query)

def get_page_views_1(client, area):
    kusto_query_1 = open(path_mbr.query_path_1,'r').read().format(area)
    kusto_database_1 = 'DevRelWorkArea'
    result_1 = query_kusto(client, kusto_database_1, kusto_query_1)
    df_1 = dataframe_from_result_table(result_1.primary_results[0])
    return df_1

def get_page_views_2(client, area):
    kusto_query_2 = open(path_mbr.query_path_2,'r').read().format(area)
    kusto_database_2 = 'DevRelWorkArea'
    result_2 = query_kusto(client, kusto_database_2, kusto_query_2)
    df_2 = dataframe_from_result_table(result_2.primary_results[0])
    return df_2

def get_page_views_3(client, area):
    kusto_query_3 = open(path_mbr.query_path_3,'r').read().format(area)
    kusto_database_3 = 'DevRelWorkArea'
    result_3 = query_kusto(client, kusto_database_3, kusto_query_3)
    df_3 = dataframe_from_result_table(result_3.primary_results[0])
    return df_3

def get_page_views_4(client, area):
    kusto_query_4 = open(path_mbr.query_path_4,'r').read().format(area)
    kusto_database_4 = 'DevRelWorkArea'
    result_4 = query_kusto(client, kusto_database_4, kusto_query_4)
    df_4 = dataframe_from_result_table(result_4.primary_results[0])
    return df_4

def get_page_views_5(client, area):
    kusto_query_5 = open(path_mbr.query_path_5,'r').read().format(area)
    kusto_database_5 = 'DevRelWorkArea'
    result_5 = query_kusto(client, kusto_database_5, kusto_query_5)
    df_5 = dataframe_from_result_table(result_5.primary_results[0])
    return df_5

def get_page_views_6(client, area):
    kusto_query_6 = open(path_mbr.query_path_6,'r').read().format(area)
    kusto_database_6 = 'DevRelWorkArea'
    result_6 = query_kusto(client, kusto_database_6, kusto_query_6)
    df_6 = dataframe_from_result_table(result_6.primary_results[0])
    return df_6

def get_page_views_7(client, area):
    kusto_query_7 = open(path_mbr.query_path_7,'r').read().format(area)
    kusto_database_7 = 'DevRelWorkArea'
    result_7 = query_kusto(client, kusto_database_7, kusto_query_7)
    df_7 = dataframe_from_result_table(result_7.primary_results[0])
    return df_7

def get_page_views_8(client, area):
    kusto_query_8 = open(path_mbr.query_path_8,'r').read().format(area)
    kusto_database_8 = 'DevRelWorkArea'
    result_8 = query_kusto(client, kusto_database_8, kusto_query_8)
    df_8 = dataframe_from_result_table(result_8.primary_results[0])
    return df_8

def get_page_views_9(client, area):
    kusto_query_9 = open(path_mbr.query_path_9,'r').read().format(area)
    kusto_database_9 = 'DevRelWorkArea'
    result_9 = query_kusto(client, kusto_database_9, kusto_query_9)
    df_9 = dataframe_from_result_table(result_9.primary_results[0])
    return df_9

def get_page_views_10(client, area):
    kusto_query_10 = open(path_mbr.query_path_10,'r').read().format(area)
    kusto_database_10 = 'DevRelWorkArea'
    result_10 = query_kusto(client, kusto_database_10, kusto_query_10)
    df_10 = dataframe_from_result_table(result_10.primary_results[0])
    return df_10

def get_page_views_11(client, area):
    kusto_query_11 = open(path_mbr.query_path_11,'r').read().format(area)
    kusto_database_11 = 'DevRelWorkArea'
    result_11 = query_kusto(client, kusto_database_11, kusto_query_11)
    df_11 = dataframe_from_result_table(result_11.primary_results[0])
    return df_11

def get_page_views_12(client, area):
    kusto_query_12 = open(path_mbr.query_path_12,'r').read().format(area)
    kusto_database_12 = 'DevRelWorkArea'
    result_12 = query_kusto(client, kusto_database_12, kusto_query_12)
    df_12 = dataframe_from_result_table(result_12.primary_results[0])
    return df_12

def get_page_views_13(client, area):
    kusto_query_13 = open(path_mbr.query_path_13,'r').read().format(area)
    kusto_database_13 = 'DevRelWorkArea'
    result_13 = query_kusto(client, kusto_database_13, kusto_query_13)
    df_13 = dataframe_from_result_table(result_13.primary_results[0])
    return df_13

def get_page_views_14(client, area):
    kusto_query_14 = open(path_mbr.query_path_14,'r').read().format(area)
    kusto_database_14 = 'DevRelWorkArea'
    result_14 = query_kusto(client, kusto_database_14, kusto_query_14)
    df_14 = dataframe_from_result_table(result_14.primary_results[0])
    return df_14

# def get_page_views_15(client, area):
#     kusto_query_15 = "database('DevRelWorkArea').['kpi-azure-impact'] " \
#                      "| where Area == '" + area + "' " \
#                      "| where Month >= startofmonth(now(),-14) and Month < startofmonth(now()) " \
#                      "| summarize val = avg(site_activated) by Month " \
#                      "| order by Month desc "
#     kusto_database_15 = 'DevRelWorkArea'
#     result_15 = query_kusto(client, kusto_database_15, kusto_query_15)
#     df_15 = dataframe_from_result_table(result_15.primary_results[0])
#     return df_15
#
# def get_page_views_16(client, area):
#     kusto_query_16 = "database('DevRelWorkArea').['kpi-azure-impact'] " \
#                      "| where Area == '" + area + "' " \
#                      "| where Month >= startofmonth(now(),-14) and Month < startofmonth(now()) " \
#                      "| summarize val = avg(site_paid) by Month " \
#                      "| order by Month desc "
#     kusto_database_16 = 'DevRelWorkArea'
#     result_16 = query_kusto(client, kusto_database_16, kusto_query_16)
#     df_16 = dataframe_from_result_table(result_16.primary_results[0])
#     return df_16
#
# def get_page_views_17(client, area):
#     kusto_query_17 = "database('DevRelWorkArea').['kpi-azure-impact'] " \
#                      "| where Area == '" + area + "' " \
#                      "| where Month >= startofmonth(now(),-14) and Month < startofmonth(now()) " \
#                      "| summarize val = avg(site_before_signup) by Month " \
#                      "| order by Month desc "
#     kusto_database_17 = 'DevRelWorkArea'
#     result_17 = query_kusto(client, kusto_database_17, kusto_query_17)
#     df_17 = dataframe_from_result_table(result_17.primary_results[0])
#     return df_17

# Query Kusto
cga_cluster = 'https://cgadataout.kusto.windows.net'
cga_client = authenticate_kusto(cga_cluster)
ls=[cga_client, "Documentation & Reference"]

current_1 = get_page_views_1(ls[0], ls[1])
current_2 = get_page_views_2(ls[0], ls[1])
current_3 = get_page_views_3(ls[0], ls[1])
current_4 = get_page_views_4(ls[0], ls[1])
current_5 = get_page_views_5(ls[0], ls[1])
current_6 = get_page_views_6(ls[0], ls[1])
current_7 = get_page_views_7(ls[0], ls[1])
current_8 = get_page_views_8(ls[0], ls[1])
current_9 = get_page_views_9(ls[0], ls[1])
current_10 = get_page_views_10(ls[0], ls[1])
current_11 = get_page_views_11(ls[0], ls[1])
current_12 = get_page_views_12(ls[0], ls[1])
current_13 = get_page_views_13(ls[0], ls[1])
current_14 = get_page_views_14(ls[0], ls[1])
# current_15 = get_page_views_15(ls[0], ls[1])
# current_16 = get_page_views_16(ls[0], ls[1])
# current_17 = get_page_views_17(ls[0], ls[1])

# Change values in Excel
wb_obj = openpyxl.load_workbook(path_mbr.path_mbr)
sheet_x = wb_obj['CONTENT']

i = 0
j = 0
x = 9
y = 23
for i in range(14):
    sheet_x.cell(x, y, current_1.iloc[i][j]) #9行表头
    sheet_x.cell(x + 1, y, current_1.iloc[i][j + 1]) #10行
    sheet_x.cell(x + 2, y, current_2.iloc[i][j + 1]) #11行
    sheet_x.cell(x + 3, y, current_3.iloc[i][j + 1]) #12行
    sheet_x.cell(x + 4, y, current_4.iloc[i][j + 1]) #13行
    sheet_x.cell(x + 5, y, current_5.iloc[i][j + 1]) #14行
    sheet_x.cell(x + 7, y, current_6.iloc[i][j + 1]) #16
    sheet_x.cell(x + 8, y, current_7.iloc[i][j + 1]) #17
    sheet_x.cell(x + 9, y, current_8.iloc[i][j + 1]) #18
    sheet_x.cell(x + 10, y, current_9.iloc[i][j + 1]) #19
    sheet_x.cell(x + 11, y, current_10.iloc[i][j + 1]) #20
    sheet_x.cell(x + 13, y, current_11.iloc[i][j + 1]) #22
    sheet_x.cell(x + 14, y, current_12.iloc[i][j + 1]) #23
    sheet_x.cell(x + 15, y, current_13.iloc[i][j + 1]) #24
    sheet_x.cell(x + 16, y, current_14.iloc[i][j + 1]) #25
    y += -1
    i += 1

wb_obj.save(path_mbr.path_mbr)