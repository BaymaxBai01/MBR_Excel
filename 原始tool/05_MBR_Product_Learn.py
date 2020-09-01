from azure.kusto.data.request import KustoClient, KustoConnectionStringBuilder
from azure.kusto.data.helpers import dataframe_from_result_table
import openpyxl
import path_mbr

def authenticate_kusto(kusto_cluster):
    tenant_id = '72f988bf-86f1-41af-91ab-2d7cd011db47'
    KCSB = KustoConnectionStringBuilder.with_aad_device_authentication(kusto_cluster)
    KCSB.authority_id = tenant_id
    return KustoClient(KCSB)

def query_kusto(client, database, query):
    return client.execute(database, query)

def get_page_views_1(client, area):
    kusto_query_1 = open(path_mbr.query_path_product_1,'r').read().format(area)
    kusto_database_1 = 'DevRelWorkArea'
    result_1 = query_kusto(client, kusto_database_1, kusto_query_1)
    df_1 = dataframe_from_result_table(result_1.primary_results[0])
    return df_1

def get_page_views_2(client, area):
    kusto_query_2 = open(path_mbr.query_path_product_2_learn, 'r').read().format(area)
    kusto_database_2 = 'DevRelWorkArea'
    result_2 = query_kusto(client, kusto_database_2, kusto_query_2)
    df_2 = dataframe_from_result_table(result_2.primary_results[0])
    return df_2

def get_page_views_3(client, area):
    kusto_query_3 = open(path_mbr.query_path_product_3_learn, 'r').read().format(area)
    kusto_database_3 = 'DevRelWorkArea'
    result_3 = query_kusto(client, kusto_database_3, kusto_query_3)
    df_3 = dataframe_from_result_table(result_3.primary_results[0])
    return df_3

def get_page_views_4(client, area):
    kusto_query_4 = open(path_mbr.query_path_product_4_learn, 'r').read().format(area)
    kusto_database_4 = 'DevRelWorkArea'
    result_4 = query_kusto(client, kusto_database_4, kusto_query_4)
    df_4 = dataframe_from_result_table(result_4.primary_results[0])
    return df_4

def get_page_views_5(client, area):
    kusto_query_5 = open(path_mbr.query_path_product_5_learn,'r').read().format(area)
    kusto_database_5 = 'DevRelWorkArea'
    result_5 = query_kusto(client, kusto_database_5, kusto_query_5)
    df_5 = dataframe_from_result_table(result_5.primary_results[0])
    return df_5

# def get_page_views_7(client, area):
#     kusto_query_7 = "database('DevRelWorkArea').['kpi-deploy-to-azure'] " \
#                     "| where Area == '" + area + "' "\
#                     "| where Month >= startofmonth(now(),-14) and Month < startofmonth(now()) " \
#                     "| summarize val = avg(uv_clicked) by Month "\
#                     "| order by Month desc "
#     kusto_database_7 = 'DevRelWorkArea'
#     result_7 = query_kusto(client, kusto_database_7, kusto_query_7)
#     df_7 = dataframe_from_result_table(result_7.primary_results[0])
#     return df_7

# Query Kusto
cga_cluster = 'https://cgadataout.kusto.windows.net'
cga_client = authenticate_kusto(cga_cluster)
ls=[cga_client, "Learn"]

current_1 = get_page_views_1(ls[0], ls[1])
current_2 = get_page_views_2(ls[0], ls[1])
current_3 = get_page_views_3(ls[0], ls[1])
current_4 = get_page_views_4(ls[0], ls[1])
current_5 = get_page_views_5(ls[0], ls[1])
# current_7 = get_page_views_7(ls[0], ls[1])

# Change values in Excel
wb_obj = openpyxl.load_workbook(path_mbr.path_mbr)
sheet_x=wb_obj['PRODUCT']

i = 0
j = 0
x = 27
y = 23
for i in range(14):
    sheet_x.cell(x + 1, y, current_1.iloc[i][j + 1]) #28行
    sheet_x.cell(x + 2, y, current_2.iloc[i][j + 1]) #29行
    sheet_x.cell(x + 3, y, current_3.iloc[i][j + 1]) #30行
    sheet_x.cell(x + 4, y, current_4.iloc[i][j + 1]) #31行
    sheet_x.cell(x + 6, y, current_5.iloc[i][j + 1]) #33行
    y += -1
    i += 1

wb_obj.save(path_mbr.path_mbr)